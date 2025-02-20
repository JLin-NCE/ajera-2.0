import pandas as pd
import json
import os
import sys
import time
import requests
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo  # For PST/PDT (Los Angeles)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
import re

class ProjectFetcher:
    def __init__(self):
        self.API_URL = "https://ajera.com/V006275/AjeraAPI.ashx?ew0KICAiQ2xpZW50SUQiOiA2Mjc1LA0KICAiRGF0YWJhc2VJRCI6IDEzMDI5LA0KICAiSXNTYW1wbGVEYXRhIjogZmFsc2UNCn0%3d"
        self.CREDENTIALS = {
            "USERNAME": "powerquerynv",
            "PASSWORD": "Testt1m3!!123"
        }
        
        # File paths
        self.FILES = {
            "NEW": "new_data.xlsx",
            "PREVIOUS": "previous_data.xlsx",
            "DEBUG": "phase_debug.log"
        }
        
        # Set up weekly directory structure using current LA time
        self.BACKLOG_DIR = "backlog"
        self.current_date = datetime.now(ZoneInfo("America/Los_Angeles"))
        self.week_start = self.current_date - timedelta(days=self.current_date.weekday())
        self.week_dir = os.path.join(self.BACKLOG_DIR, self.week_start.strftime("%Y-%m-%d"))
        
        # Create directories if they do not exist
        os.makedirs(self.BACKLOG_DIR, exist_ok=True)
        os.makedirs(self.week_dir, exist_ok=True)
        
        self.BATCH_SIZE = 25
        self.session_token = None
        self.retry_count = 3
        self.retry_delay = 1
        
        # Added attribute to force conversion to PDT
        self.force_pdt = True  # NEW FUNCTIONALITY 1

        # NEW: Initialize a list to store debug log entries (timestamp, message)
        self.debug_entries = []
        
        # Initialize debug log (text file)
        with open(self.FILES["DEBUG"], 'w') as f:
            init_msg = f"Debug Log Started: {datetime.now(ZoneInfo('America/Los_Angeles'))}\n" + ("-" * 80 + "\n")
            f.write(init_msg)
        
        # Initialize data fetched time (we’ll store a fresh LA time each run)
        self.data_fetched_time = datetime.now(ZoneInfo("America/Los_Angeles"))

    def log_debug(self, message):
        """Write debug information to the log file and store in memory."""
        timestamp = datetime.now(ZoneInfo("America/Los_Angeles")).strftime("%Y-%m-%d %H:%M:%S")
        log_line = f"[{timestamp}] {message}"
        self.debug_entries.append((timestamp, message))
        with open(self.FILES["DEBUG"], 'a') as f:
            f.write(log_line + "\n")

    def backup_current_data(self):
        """Backup the current new_data file before fetching new data."""
        try:
            if os.path.exists(self.FILES["NEW"]):
                print("\n📂 Backing up current data...")
                if os.path.exists(self.FILES["PREVIOUS"]):
                    os.remove(self.FILES["PREVIOUS"])
                os.rename(self.FILES["NEW"], self.FILES["PREVIOUS"])
                print("✅ Current data backed up as previous_data.xlsx")
                self.log_debug("Backed up new_data.xlsx to previous_data.xlsx")
            else:
                print("\nℹ️ No existing data file to backup")
                self.log_debug("No existing new_data.xlsx to backup")
        except Exception as e:
            print(f"❌ Backup failed: {str(e)}")
            self.log_debug(f"Backup failed: {str(e)}")
            raise

    def create_session(self):
        """Create an API session and retrieve the session token."""
        try:
            print("\n🔑 Creating API session...")
            self.log_debug("Creating API session...")
            
            payload = {
                "Method": "CreateAPISession",
                "Username": self.CREDENTIALS["USERNAME"],
                "Password": self.CREDENTIALS["PASSWORD"],
                "APIVersion": 2
            }
            
            response = requests.post(self.API_URL, json=payload)
            response.raise_for_status()
            
            data = response.json()
            if "Errors" in data and data["Errors"]:
                error_msg = data["Errors"][0].get("ErrorMessage")
                self.log_debug(f"API Error: {error_msg}")
                raise Exception(f"API Error: {error_msg}")
            
            self.session_token = data["Content"]["SessionToken"]
            print("✅ Session created successfully")
            self.log_debug("Session created successfully")
            return self.session_token
            
        except Exception as e:
            print(f"❌ Session creation failed: {str(e)}")
            self.log_debug(f"Session creation failed: {str(e)}")
            raise

    def parse_and_convert_to_la(self, date_str, force_pdt=False):
        """
        Parse a date string like:
          "2023-03-13T15:21:11.190 GMT-5 (Eastern Standard Time)"
        and convert it to Los Angeles time (PST/PDT). 
        NEW: If force_pdt is True, the returned string will always end with 'PDT'.
        If parsing fails, return the original string.
        """
        if not date_str:
            return date_str

        # Regex to capture the datetime and the offset portion
        pattern = r"^(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d+)\s+GMT([+-]\d+)"
        match = re.search(pattern, date_str)
        if not match:
            return date_str
        
        dt_part = match.group(1)   # e.g. "2023-03-13T15:21:11.190"
        offset_str = match.group(2)  # e.g. "-5"

        try:
            dt_obj = datetime.fromisoformat(dt_part)
        except ValueError:
            return date_str  # fallback if we can't parse

        try:
            offset_hours = int(offset_str)
        except ValueError:
            return date_str

        offset = timezone(timedelta(hours=offset_hours))
        dt_with_offset = dt_obj.replace(tzinfo=offset)

        # Convert to Los Angeles time
        la_time = dt_with_offset.astimezone(ZoneInfo("America/Los_Angeles"))
        if force_pdt:
            # Force the timezone abbreviation to "PDT" regardless of actual DST status.
            formatted_time = la_time.strftime("%Y-%m-%d %H:%M:%S") + " PDT"
        else:
            formatted_time = la_time.strftime("%Y-%m-%d %H:%M:%S %Z")
        return formatted_time

    def fetch_projects(self):
        """Fetch all active projects and record the fetch time in LA (PDT/PST)."""
        try:
            print("\n📋 Fetching active projects...")
            self.log_debug("Fetching active projects...")
            
            payload = {
                "Method": "ListProjects",
                "SessionToken": self.session_token,
                "MethodArguments": {
                    "FilterByStatus": ["Active"]
                }
            }
            
            response = requests.post(self.API_URL, json=payload)
            response.raise_for_status()
            
            data = response.json()
            if "Errors" in data and data["Errors"]:
                error_msg = data["Errors"][0].get("ErrorMessage")
                self.log_debug(f"API Error: {error_msg}")
                raise Exception(f"API Error: {error_msg}")
            
            projects = data["Content"].get("Projects", [])
            print(f"✅ Retrieved {len(projects)} projects")
            self.log_debug(f"Retrieved {len(projects)} projects")
            
            return projects
            
        except Exception as e:
            print(f"❌ Project fetch failed: {str(e)}")
            self.log_debug(f"Project fetch failed: {str(e)}")
            raise

    def fetch_project_details(self, project_key, attempt=1):
        """Fetch detailed information for a single project with retry logic, converting times to LA (forced to PDT if set)."""
        try:
            self.log_debug(f"Fetching details for project {project_key}")
            
            payload = {
                "Method": "GetProjects",
                "SessionToken": self.session_token,
                "MethodArguments": {
                    "RequestedProjects": [project_key],
                    "IncludePhases": True
                }
            }
            
            response = requests.post(self.API_URL, json=payload)
            response.raise_for_status()
            
            data = response.json()
            if "Errors" in data and data["Errors"]:
                error_msg = data["Errors"][0].get("ErrorMessage")
                self.log_debug(f"API Error for project {project_key}: {error_msg}")
                raise Exception(f"API Error: {error_msg}")
            
            project = data["Content"].get("Projects", [])[0]
            phases = data["Content"].get("Phases", [])

            # Convert project's LastModifiedDate using forced PDT if set
            if "LastModifiedDate" in project:
                project["LastModifiedDate"] = self.parse_and_convert_to_la(project["LastModifiedDate"], force_pdt=self.force_pdt)

            project["Date Fetched"] = self.data_fetched_time.strftime("%Y-%m-%d %H:%M:%S")

            # Convert each phase's LastModifiedDate using forced PDT if set
            for phase in phases:
                if "LastModifiedDate" in phase:
                    phase["LastModifiedDate"] = self.parse_and_convert_to_la(phase["LastModifiedDate"], force_pdt=self.force_pdt)
                phase["Date Fetched"] = self.data_fetched_time.strftime("%Y-%m-%d %H:%M:%S")

            project["Phases"] = phases

            self.log_debug(f"Project {project_key}:")
            self.log_debug(f"  - Description: {project.get('Description', 'N/A')}")
            self.log_debug(f"  - Status: {project.get('Status', 'N/A')}")
            self.log_debug(f"  - Total Contract Amount: {project.get('TotalContractAmount', 0)}")
            self.log_debug(f"  - Phase count: {len(phases)}")
            
            return project
            
        except Exception as e:
            if attempt < self.retry_count:
                self.log_debug(f"Retry {attempt} for project {project_key}")
                time.sleep(self.retry_delay * attempt)
                return self.fetch_project_details(project_key, attempt + 1)
            self.log_debug(f"Failed to fetch details for project {project_key}: {str(e)}")
            return None

    def process_projects_in_batches(self, projects):
        """Process projects in batches for efficiency."""
        try:
            if not projects:
                return []

            total_projects = len(projects)
            processed_data = []
            
            print(f"\n🔄 Processing {total_projects} projects in batches of {self.BATCH_SIZE}...")
            self.log_debug(f"Processing {total_projects} projects in batches of {self.BATCH_SIZE}")
            
            for i in range(0, total_projects, self.BATCH_SIZE):
                batch = projects[i:i + self.BATCH_SIZE]
                batch_number = (i // self.BATCH_SIZE) + 1
                total_batches = (total_projects + self.BATCH_SIZE - 1) // self.BATCH_SIZE
                
                print(f"\n📦 Processing batch {batch_number} of {total_batches} ({len(batch)} projects)")
                self.log_debug(f"Processing batch {batch_number} of {total_batches} ({len(batch)} projects)")
                
                for idx, project in enumerate(batch, 1):
                    project_key = project["ProjectKey"]
                    progress = ((i + idx) / total_projects) * 100
                    print(f"  Processing project {project_key}... ({progress:.1f}%)")
                    
                    project_details = self.fetch_project_details(project_key)
                    if project_details:
                        processed_data.append(project_details)
                    
                if i + self.BATCH_SIZE < total_projects:
                    time.sleep(0.5)
            
            print(f"\n✅ Successfully processed {len(processed_data)} projects")
            self.log_debug(f"Successfully processed {len(processed_data)} projects")
            return processed_data
            
        except Exception as e:
            print(f"❌ Batch processing failed: {str(e)}")
            self.log_debug(f"Batch processing failed: {str(e)}")
            raise

    def save_to_excel(self, data):
        """
        Save project and phase data to Excel. 
        We now pull the 'LastModifiedDate' and 'Date Fetched' directly from each project/phase.
        """
        try:
            print("\n💾 Saving data to Excel...")
            self.log_debug("Saving data to Excel...")
            
            wb = Workbook()
            
            # Projects sheet
            ws_projects = wb.active
            ws_projects.title = "Projects"
            
            headers = [
                "Project ID", 
                "Description",
                "Total Contract Amount",
                "Status",
                "Last Modified",
                "Date Fetched"
            ]
            ws_projects.append(headers)
            for cell in ws_projects[1]:
                cell.font = Font(bold=True)
            
            project_count = 0
            for project in data:
                ws_projects.append([
                    project.get("ID"),
                    project.get("Description"),
                    project.get("TotalContractAmount", 0),
                    project.get("Status"),
                    project.get("LastModifiedDate"),
                    project.get("Date Fetched")
                ])
                project_count += 1
            
            # Phases sheet
            ws_phases = wb.create_sheet("Phases")
            headers = [
                "Project ID",
                "Phase ID",
                "Description",
                "Total Contract Amount",
                "Status",
                "Last Modified",
                "Date Fetched"
            ]
            ws_phases.append(headers)
            for cell in ws_phases[1]:
                cell.font = Font(bold=True)
            
            phase_count = 0
            for project in data:
                project_id = project.get("ID")
                for phase in project.get("Phases", []):
                    ws_phases.append([
                        project_id,
                        phase.get("ID"),
                        phase.get("Description"),
                        phase.get("TotalContractAmount", 0),
                        phase.get("Status"),
                        phase.get("LastModifiedDate"),
                        phase.get("Date Fetched")
                    ])
                    phase_count += 1
            
            wb.save(self.FILES["NEW"])
            print(f"✅ Saved {project_count} projects and {phase_count} phases to {self.FILES['NEW']}")
            self.log_debug(f"Saved {project_count} projects and {phase_count} phases to {self.FILES['NEW']}")
            
            return project_count, phase_count
            
        except Exception as e:
            print(f"❌ Save failed: {str(e)}")
            self.log_debug(f"Save failed: {str(e)}")
            raise

    def compare_contract_amounts(self):
        """
        Compare TotalContractAmount between new and previous data.
        NEW FUNCTIONALITY 2: Also capture and compare the "Last Modified" values.
        Always returns a changes dictionary (even if empty) so that a report is generated.
        """
        try:
            if not os.path.exists(self.FILES["PREVIOUS"]):
                print("\nℹ️ No previous data to compare")
                self.log_debug("No previous data to compare")
                return {
                    "projects": [],
                    "phases": [],
                    "validation": {
                        "missing_phases": [],
                        "phase_status_changes": []
                    }
                }
            
            print("\n🔍 Comparing contract amounts...")
            self.log_debug("Comparing contract amounts...")
            
            old_wb = load_workbook(self.FILES["PREVIOUS"])
            new_wb = load_workbook(self.FILES["NEW"])
            
            changes = {
                "projects": [],
                "phases": [],
                "validation": {
                    "missing_phases": [],
                    "phase_status_changes": []
                }
            }
            
            old_projects = {
                row[0].value: {
                    'amount': row[2].value,
                    'description': row[1].value,
                    'status': row[3].value,
                    'last_modified': row[4].value
                }
                for row in old_wb["Projects"].iter_rows(min_row=2)
                if row[0].value
            }
            
            new_projects = {
                row[0].value: {
                    'amount': row[2].value,
                    'description': row[1].value,
                    'status': row[3].value,
                    'last_modified': row[4].value
                }
                for row in new_wb["Projects"].iter_rows(min_row=2)
                if row[0].value
            }
            
            for project_id in set(old_projects) | set(new_projects):
                old_data = old_projects.get(project_id, {'amount': 0, 'description': 'N/A', 'status': 'N/A', 'last_modified': 'N/A'})
                new_data = new_projects.get(project_id, {'amount': 0, 'description': 'N/A', 'status': 'N/A', 'last_modified': 'N/A'})
                
                old_amount = old_data['amount'] or 0
                new_amount = new_data['amount'] or 0
                
                if old_amount != new_amount:
                    changes["projects"].append({
                        "ID": project_id,
                        "Description": new_data['description'],
                        "Status": new_data['status'],
                        "Old Amount": old_amount,
                        "New Amount": new_amount,
                        "Change": new_amount - old_amount,
                        "Old Last Modified": old_data.get('last_modified'),
                        "New Last Modified": new_data.get('last_modified')
                    })
                    self.log_debug(f"Project Change Detected: {project_id} {new_data['description']} from {old_amount} to {new_amount}")
            
            old_phases = {
                (row[0].value, row[1].value): {
                    'amount': row[3].value,
                    'description': row[2].value,
                    'status': row[4].value,
                    'last_modified': row[5].value
                }
                for row in old_wb["Phases"].iter_rows(min_row=2)
                if row[0].value and row[1].value
            }
            
            new_phases = {
                (row[0].value, row[1].value): {
                    'amount': row[3].value,
                    'description': row[2].value,
                    'status': row[4].value,
                    'last_modified': row[5].value
                }
                for row in new_wb["Phases"].iter_rows(min_row=2)
                if row[0].value and row[1].value
            }
            
            for phase_key in set(old_phases) | set(new_phases):
                project_id, phase_id = phase_key
                old_data = old_phases.get(phase_key, {'amount': 0, 'description': 'N/A', 'status': 'N/A', 'last_modified': 'N/A'})
                new_data = new_phases.get(phase_key, {'amount': 0, 'description': 'N/A', 'status': 'N/A', 'last_modified': 'N/A'})
                
                if phase_key in old_phases and phase_key not in new_phases:
                    changes["validation"]["missing_phases"].append({
                        "Project ID": project_id,
                        "Phase ID": phase_id,
                        "Description": old_data['description'],
                        "Status": old_data['status'],
                        "Amount": old_data['amount']
                    })
                    self.log_debug(f"Missing Phase: Project {project_id} Phase {phase_id}")
                
                if (phase_key in old_phases and phase_key in new_phases and 
                    old_data['status'] != new_data['status']):
                    changes["validation"]["phase_status_changes"].append({
                        "Project ID": project_id,
                        "Phase ID": phase_id,
                        "Description": new_data['description'],
                        "Old Status": old_data['status'],
                        "New Status": new_data['status']
                    })
                
                old_amount = old_data['amount'] or 0
                new_amount = new_data['amount'] or 0
                if old_amount != new_amount:
                    changes["phases"].append({
                        "Project ID": project_id,
                        "Phase ID": phase_id,
                        "Description": new_data['description'],
                        "Status": new_data['status'],
                        "Old Amount": old_amount,
                        "New Amount": new_amount,
                        "Change": new_amount - old_amount,
                        "Old Last Modified": old_data.get('last_modified'),
                        "New Last Modified": new_data.get('last_modified')
                    })
                    self.log_debug(f"Phase Change: Project {project_id} Phase {phase_id} from {old_amount} to {new_amount}")
            
            total_changes = len(changes["projects"]) + len(changes["phases"])
            print(f"✅ Found {total_changes} total changes:")
            print(f"  - Project changes: {len(changes['projects'])}")
            print(f"  - Phase changes: {len(changes['phases'])}")
            self.log_debug(f"Comparison Summary: {len(changes['projects'])} project changes, {len(changes['phases'])} phase changes")
            
            return changes
            
        except Exception as e:
            print(f"❌ Comparison failed: {str(e)}")
            self.log_debug(f"Comparison failed: {str(e)}")
            raise

    def save_comparison_report(self, changes, filename):
        """
        Save the comparison report to an Excel file with three sheets:
          1. Project Changes
          2. Phase Changes
          3. Filtered Phase Changes (only for projects with overall changes)
        NEW FUNCTIONALITY 2: The report now includes two extra columns for the "Last Modified" fields
                              from both previous_data and new_data, plus a new "Flag" column.
        """
        try:
            wb = Workbook()
            
            # --- Sheet 1: Project Changes with Flag ---
            ws_projects = wb.active
            ws_projects.title = "Project Changes"
            
            proj_headers = [
                "Project ID",
                "Description",
                "Status",
                "Old Amount",
                "New Amount",
                "Change",
                "Previous Last Modified",
                "New Last Modified",
                "Flag",             # NEW: Flag column
                "Change %"
            ]
            ws_projects.append(proj_headers)
            for cell in ws_projects[1]:
                cell.font = Font(bold=True)
            
            for change in changes["projects"]:
                old_amount = change["Old Amount"] or 0
                new_amount = change["New Amount"] or 0
                diff = change["Change"]
                # Compute change percent
                change_percent = ((diff / old_amount) * 100) if old_amount != 0 else 100
                # Compute flag based on change
                if old_amount == 0:
                    flag = "New" if new_amount > 0 else ""
                else:
                    if diff > 0:
                        flag = "Increase"
                    elif diff < 0:
                        flag = "Decrease"
                    else:
                        flag = ""
                if old_amount != 0 and abs((diff / old_amount) * 100) >= 10:
                    flag = (flag + " (Significant)").strip()
                row = [
                    change["ID"],
                    change["Description"],
                    change["Status"],
                    change["Old Amount"],
                    change["New Amount"],
                    diff,
                    change.get("Old Last Modified", ""),
                    change.get("New Last Modified", ""),
                    flag,
                    f"{change_percent:.2f}%"
                ]
                ws_projects.append(row)
                # Highlight the "Change" and "Change %" cells based on the sign
                change_cell = ws_projects.cell(row=ws_projects.max_row, column=6)
                percent_cell = ws_projects.cell(row=ws_projects.max_row, column=10)
                fill = PatternFill(
                    start_color="FFCDD2" if diff < 0 else "C8E6C9",
                    end_color="FFCDD2" if diff < 0 else "C8E6C9",
                    fill_type="solid"
                )
                change_cell.fill = fill
                percent_cell.fill = fill
            
            # --- Sheet 2: Phase Changes with Flag ---
            ws_phases = wb.create_sheet("Phase Changes")
            phase_headers = [
                "Project ID",
                "Phase ID",
                "Description",
                "Status",
                "Old Amount",
                "New Amount",
                "Change",
                "Previous Last Modified",
                "New Last Modified",
                "Flag",             # NEW: Flag column
                "Change %"
            ]
            ws_phases.append(phase_headers)
            for cell in ws_phases[1]:
                cell.font = Font(bold=True)
            
            for change in changes["phases"]:
                old_amount = change["Old Amount"] or 0
                new_amount = change["New Amount"] or 0
                diff = change["Change"]
                change_percent = ((diff / old_amount) * 100) if old_amount != 0 else 100
                if old_amount == 0:
                    flag = "New" if new_amount > 0 else ""
                else:
                    if diff > 0:
                        flag = "Increase"
                    elif diff < 0:
                        flag = "Decrease"
                    else:
                        flag = ""
                if old_amount != 0 and abs((diff / old_amount) * 100) >= 10:
                    flag = (flag + " (Significant)").strip()
                row = [
                    change["Project ID"],
                    change["Phase ID"],
                    change["Description"],
                    change["Status"],
                    change["Old Amount"],
                    change["New Amount"],
                    diff,
                    change.get("Old Last Modified", ""),
                    change.get("New Last Modified", ""),
                    flag,
                    f"{change_percent:.2f}%"
                ]
                ws_phases.append(row)
                change_cell = ws_phases.cell(row=ws_phases.max_row, column=7)
                percent_cell = ws_phases.cell(row=ws_phases.max_row, column=11)
                fill = PatternFill(
                    start_color="FFCDD2" if diff < 0 else "C8E6C9",
                    end_color="FFCDD2" if diff < 0 else "C8E6C9",
                    fill_type="solid"
                )
                change_cell.fill = fill
                percent_cell.fill = fill
            
            # --- Sheet 3: Filtered Phase Changes with Flag ---
            ws_filt = wb.create_sheet("Filtered Phase Changes")
            filt_headers = [
                "Project ID",
                "Phase ID",
                "Description",
                "Status",
                "Old Amount",
                "New Amount",
                "Change",
                "Previous Last Modified",
                "New Last Modified",
                "Flag",             # NEW: Flag column
                "Change %"
            ]
            ws_filt.append(filt_headers)
            for cell in ws_filt[1]:
                cell.font = Font(bold=True)
            
            changed_project_ids = set(item["ID"] for item in changes["projects"])
            for change in changes["phases"]:
                if change["Project ID"] in changed_project_ids:
                    old_amount = change["Old Amount"] or 0
                    new_amount = change["New Amount"] or 0
                    diff = change["Change"]
                    change_percent = ((diff / old_amount) * 100) if old_amount != 0 else 100
                    if old_amount == 0:
                        flag = "New" if new_amount > 0 else ""
                    else:
                        if diff > 0:
                            flag = "Increase"
                        elif diff < 0:
                            flag = "Decrease"
                        else:
                            flag = ""
                    if old_amount != 0 and abs((diff / old_amount) * 100) >= 10:
                        flag = (flag + " (Significant)").strip()
                    row = [
                        change["Project ID"],
                        change["Phase ID"],
                        change["Description"],
                        change["Status"],
                        change["Old Amount"],
                        change["New Amount"],
                        diff,
                        change.get("Old Last Modified", ""),
                        change.get("New Last Modified", ""),
                        flag,
                        f"{change_percent:.2f}%"
                    ]
                    ws_filt.append(row)
                    change_cell = ws_filt.cell(row=ws_filt.max_row, column=7)
                    percent_cell = ws_filt.cell(row=ws_filt.max_row, column=11)
                    fill = PatternFill(
                        start_color="FFCDD2" if diff < 0 else "C8E6C9",
                        end_color="FFCDD2" if diff < 0 else "C8E6C9",
                        fill_type="solid"
                    )
                    change_cell.fill = fill
                    percent_cell.fill = fill
            
            wb.save(filename)
            print(f"✅ Changes report saved to {filename}")
            self.log_debug(f"Changes report saved to {filename}")
            
        except Exception as e:
            print(f"❌ Report generation failed: {str(e)}")
            self.log_debug(f"Report generation failed: {str(e)}")
            raise

    def save_debug_log_excel(self, filename):
        """
        NEW FUNCTIONALITY: Save the debug log entries (stored in self.debug_entries) into an Excel file,
        making it easier to navigate the log.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Debug Log"
            headers = ["Timestamp", "Message"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for entry in self.debug_entries:
                ws.append(list(entry))
            wb.save(filename)
            print(f"✅ Debug log Excel saved to {filename}")
            self.log_debug(f"Debug log Excel saved to {filename}")
        except Exception as e:
            print(f"❌ Saving debug log Excel failed: {str(e)}")
            self.log_debug(f"Saving debug log Excel failed: {str(e)}")
            raise

    def display_changes(self, changes):
        """Display changes in a GUI window."""
        root = tk.Tk()
        root.title("Contract Amount Changes")
        root.geometry("1200x800")
        
        style = ttk.Style()
        style.configure("Heading.TLabel", font=("Arial", 12, "bold"))
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
        style.configure("Treeview", font=("Arial", 10))
        
        summary_frame = ttk.Frame(root, padding="10")
        summary_frame.pack(fill=tk.X)
        
        total_project_change = sum(c["Change"] for c in changes["projects"])
        total_phase_change = sum(c["Change"] for c in changes["phases"])
        
        summary_text = (
            f"Summary:\n"
            f"Project Changes: {len(changes['projects'])} (Total: ${total_project_change:,.2f})\n"
            f"Phase Changes: {len(changes['phases'])} (Total: ${total_phase_change:,.2f})\n"
            f"Filtered Phase Changes: {len([ph for ph in changes['phases'] if ph['Project ID'] in set(item['ID'] for item in changes['projects'])])}"
        )
        
        ttk.Label(summary_frame, text=summary_text, style="Heading.TLabel").pack(anchor="w")
        
        notebook = ttk.Notebook(root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Projects Tab
        projects_frame = ttk.Frame(notebook)
        notebook.add(projects_frame, text="Project Changes")
        projects_tree = ttk.Treeview(
            projects_frame,
            columns=("ID", "Description", "Status", "Old", "New", "Change", "Change %"),
            show="headings"
        )
        projects_tree.heading("ID", text="Project ID")
        projects_tree.heading("Description", text="Description")
        projects_tree.heading("Status", text="Status")
        projects_tree.heading("Old", text="Old Amount")
        projects_tree.heading("New", text="New Amount")
        projects_tree.heading("Change", text="Change")
        projects_tree.heading("Change %", text="Change %")
        projects_tree.column("ID", width=100)
        projects_tree.column("Description", width=300)
        projects_tree.column("Status", width=100)
        projects_tree.column("Old", width=120)
        projects_tree.column("New", width=120)
        projects_tree.column("Change", width=120)
        projects_tree.column("Change %", width=100)
        for change in changes["projects"]:
            old_amount = change["Old Amount"] or 0
            cp = ((change["Change"] / old_amount) * 100) if old_amount != 0 else 100
            projects_tree.insert("", "end", values=(
                change["ID"],
                change["Description"],
                change["Status"],
                f"${change['Old Amount']:,.2f}",
                f"${change['New Amount']:,.2f}",
                f"${change['Change']:,.2f}",
                f"{cp:.2f}%"
            ))
        projects_scroll = ttk.Scrollbar(projects_frame, orient=tk.VERTICAL, command=projects_tree.yview)
        projects_tree.configure(yscrollcommand=projects_scroll.set)
        projects_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        projects_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Phases Tab
        phases_frame = ttk.Frame(notebook)
        notebook.add(phases_frame, text="Phase Changes")
        phases_tree = ttk.Treeview(
            phases_frame,
            columns=("Project", "Phase", "Description", "Status", "Old", "New", "Change", "Change %"),
            show="headings"
        )
        phases_tree.heading("Project", text="Project ID")
        phases_tree.heading("Phase", text="Phase ID")
        phases_tree.heading("Description", text="Description")
        phases_tree.heading("Status", text="Status")
        phases_tree.heading("Old", text="Old Amount")
        phases_tree.heading("New", text="New Amount")
        phases_tree.heading("Change", text="Change")
        phases_tree.heading("Change %", text="Change %")
        phases_tree.column("Project", width=100)
        phases_tree.column("Phase", width=100)
        phases_tree.column("Description", width=250)
        phases_tree.column("Status", width=100)
        phases_tree.column("Old", width=120)
        phases_tree.column("New", width=120)
        phases_tree.column("Change", width=120)
        phases_tree.column("Change %", width=100)
        for change in changes["phases"]:
            old_amount = change["Old Amount"] or 0
            cp = ((change["Change"] / old_amount) * 100) if old_amount != 0 else 100
            phases_tree.insert("", "end", values=(
                change["Project ID"],
                change["Phase ID"],
                change["Description"],
                change["Status"],
                f"${change['Old Amount']:,.2f}",
                f"${change['New Amount']:,.2f}",
                f"${change['Change']:,.2f}",
                f"{cp:.2f}%"
            ))
        phases_scroll = ttk.Scrollbar(phases_frame, orient=tk.VERTICAL, command=phases_tree.yview)
        phases_tree.configure(yscrollcommand=phases_scroll.set)
        phases_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        phases_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        ttk.Button(root, text="Close", command=root.destroy).pack(pady=10)
        root.mainloop()

    def generate_dashboard_html(self, projects, filename="dashboard.html"):
        """
        Generate a separate HTML dashboard page that displays one project at a time.
        This serves as a visual version of the backlog (new_data) Excel sheet.
        """
        try:
            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Project Dashboard</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        .container {{ max-width: 800px; margin: auto; }}
        .project {{ border: 1px solid #ccc; padding: 20px; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
        .nav-buttons {{ margin-top: 20px; text-align: center; }}
        button {{ padding: 10px 20px; margin: 5px; }}
    </style>
</head>
<body>
<div class="container">
  <h1>Project Dashboard</h1>
  <div id="project-container">
    <!-- Project details will be inserted here -->
  </div>
  <div class="nav-buttons">
     <button id="prevBtn">Previous</button>
     <button id="nextBtn">Next</button>
  </div>
</div>
<script>
  var projects = {json.dumps(projects)};
  var currentIndex = 0;
  function displayProject(index) {{
     var proj = projects[index];
     var html = "<div class='project'>";
     html += "<h2>Project: " + (proj.Description || '') + " (ID: " + (proj.ID || '') + ")</h2>";
     html += "<p><strong>Total Contract Amount:</strong> " + (proj.TotalContractAmount || 0) + "</p>";
     html += "<p><strong>Status:</strong> " + (proj.Status || '') + "</p>";
     html += "<p><strong>Last Modified:</strong> " + (proj.LastModifiedDate || '') + "</p>";
     html += "<p><strong>Date Fetched:</strong> " + (proj["Date Fetched"] || '') + "</p>";
     if (proj.Phases && proj.Phases.length > 0) {{
         html += "<h3>Phases</h3>";
         html += "<table><tr><th>Phase ID</th><th>Description</th><th>Total Contract Amount</th><th>Status</th><th>Last Modified</th><th>Date Fetched</th></tr>";
         for (var i = 0; i < proj.Phases.length; i++) {{
             var ph = proj.Phases[i];
             html += "<tr>";
             html += "<td>" + (ph.ID || '') + "</td>";
             html += "<td>" + (ph.Description || '') + "</td>";
             html += "<td>" + (ph.TotalContractAmount || 0) + "</td>";
             html += "<td>" + (ph.Status || '') + "</td>";
             html += "<td>" + (ph.LastModifiedDate || '') + "</td>";
             html += "<td>" + (ph["Date Fetched"] || '') + "</td>";
             html += "</tr>";
         }}
         html += "</table>";
     }} else {{
         html += "<p>No phases available.</p>";
     }}
     html += "</div>";
     document.getElementById("project-container").innerHTML = html;
  }}
  document.getElementById("prevBtn").addEventListener("click", function() {{
     if (currentIndex > 0) {{
         currentIndex--;
         displayProject(currentIndex);
     }}
  }});
  document.getElementById("nextBtn").addEventListener("click", function() {{
     if (currentIndex < projects.length - 1) {{
         currentIndex++;
         displayProject(currentIndex);
     }}
  }});
  displayProject(currentIndex);
</script>
</body>
</html>
"""
            with open(filename, "w", encoding="utf-8") as f:
                f.write(html_content)
            print(f"✅ Dashboard HTML generated and saved to {filename}")
            self.log_debug(f"Dashboard HTML generated and saved to {filename}")
        except Exception as e:
            print(f"❌ Dashboard HTML generation failed: {str(e)}")
            self.log_debug(f"Dashboard HTML generation failed: {str(e)}")
            raise

def main():
    try:
        print("\n🚀 Starting Project Data Collection")
        fetcher = ProjectFetcher()
        
        # Step 1: Backup current data
        fetcher.backup_current_data()
        
        # Step 2: Create session and fetch new data
        fetcher.create_session()
        projects = fetcher.fetch_projects()
        
        # Step 3: Process projects in batches
        processed_data = fetcher.process_projects_in_batches(projects)
        
        # Step 4: Save new data
        project_count, phase_count = fetcher.save_to_excel(processed_data)
        
        # Step 5: Compare contract amounts (including Last Modified fields)
        changes = fetcher.compare_contract_amounts()
        
        print("\n✨ Process completed successfully!")
        print(f"📊 Summary:")
        print(f"  - Projects processed: {project_count}")
        print(f"  - Phases processed: {phase_count}")
        print(f"  - Project changes: {len(changes['projects'])}")
        print(f"  - Phase changes: {len(changes['phases'])}")
        print(f"  - Filtered Phase changes: {len([ph for ph in changes['phases'] if ph['Project ID'] in set(item['ID'] for item in changes['projects'])])}")
        
        # Generate the changes report (Excel file in the backlog directory)
        current_time = datetime.now(ZoneInfo("America/Los_Angeles"))
        date_str = current_time.strftime("%b-%d-%Y")
        time_str = current_time.strftime("%I-%M-%p")
        report_file = os.path.join(fetcher.week_dir, f"changes_report_{date_str}_at_{time_str}.xlsx")
        fetcher.save_comparison_report(changes, report_file)
        
        # Save the debug log into an Excel file for easier navigation
        debug_log_file = os.path.join(fetcher.week_dir, f"debug_log_{date_str}_at_{time_str}.xlsx")
        fetcher.save_debug_log_excel(debug_log_file)
        
        # Display changes in a GUI window
        fetcher.display_changes(changes)
        
        # Generate a separate HTML dashboard page
        dashboard_file = os.path.join(fetcher.week_dir, f"dashboard_{date_str}_at_{time_str}.html")
        fetcher.generate_dashboard_html(processed_data, dashboard_file)
        
    except Exception as e:
        print(f"\n🚨 Critical error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
